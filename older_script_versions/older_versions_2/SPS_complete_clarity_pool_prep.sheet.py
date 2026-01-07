#!/usr/bin/env python3
# -*- coding: utf-8 -*-


# USAGE: SPS_complete_clarity_pool_prep_sheet.py 
    
import pandas as pd
import numpy as np
import sys
import os
import shutil
from datetime import datetime
from pathlib import Path
# from collections import defaultdict
import openpyxl
import re
# import xlrd
# from xlutils.copy import copy
# import xlwt


##########################
##########################
def findPoolPrepFile(crnt_dir):
    
    clarity_pool_prep_files = []

    # search through current directory for Clarity Pool Prep file
    # format is PoolingPrep_XX-XXXXXXXX.xlsx
    for file in os.listdir(crnt_dir):
        if file.startswith('PoolingPrep_'):
            
            # add any files matching pattern to list
            clarity_pool_prep_files.append(file)
          
    # abort script if no pool creation files are found or 
    # if >1 pool creation file found        
    if len(clarity_pool_prep_files) == 0:
        print('\n\nCould not find Clarity pool prep file, e.g. PoolingPrep_XX-XXXXX.xslx   Aborting script\n\n')
        sys.exit()
    elif len(clarity_pool_prep_files) > 1:
        print(f'\n\nMultiple Clarity pool prep files found\n\n{clarity_pool_prep_files}\n\nAborting script\n\n')
        sys.exit()

    pool_file = clarity_pool_prep_files[0]

    # return Clarity pool creation file name
    return pool_file
##########################
##########################


##########################
##########################
def getConstantsFromXlsx():
    wb = openpyxl.load_workbook(
        filename='assign_pool_number_sheet.xlsx', data_only=True)

    sheet = wb['Pooling_tool']

    min_tran_vol = sheet['Q2'].value

    max_conc_vol = sheet['Q4'].value

    max_dilut_vol = sheet['Q6'].value

    target_pool_mass = sheet['Q8'].value

    concentrate_switch = sheet['Q14'].value

    # make list of all constants imported from pooling tool .xlsx
    const_list = [target_pool_mass, max_conc_vol,
                  max_dilut_vol, min_tran_vol, concentrate_switch]

    # abort script if any constants are missing
    if all(v for v in const_list) == False:
        print("\n\nPooling tool .xlsx file is missing a constant, e.g. target library concentration, min/max transfer volume. Aborting process.")
        sys.exit()

    return target_pool_mass, max_conc_vol, max_dilut_vol, min_tran_vol, concentrate_switch
##########################
##########################


##########################
##########################
def checkRedoPlates(x_df):
    
    plate_list = x_df['Pool_source_plate'].unique().tolist()
    
    for p in plate_list:
        
        if re.search("\.\d$",p):
            name = p.split('.')
            
            if name[0] in plate_list:
                
                
                orig_index = x_df.loc[x_df['Pool_source_plate'] == name[0], 'Pool_Illumina_index_set'].values[0]
                
                redo_index = x_df.loc[x_df['Pool_source_plate'] == p, 'Pool_Illumina_index_set'].values[0]
                
                orig_pool = x_df.loc[x_df['Pool_source_plate'] == name[0], "Pool_number"].values[0]
                
                redo_pool = x_df.loc[x_df['Pool_source_plate'] == p, "Pool_number"].values[0]
                
                
                
                
                if (orig_index == redo_index) and (orig_pool == redo_pool):
                    
                    # indexAge = df[ (df['Age'] >= 20) & (df['Age'] <= 25) ].index
                    # df.drop(indexAge , inplace=True)
                    
                    redo_row = x_df[x_df['Pool_source_plate'] == p].index
                    
                    x_df.drop(redo_row, inplace=True)
                    
                elif (orig_pool != redo_pool):
                    
                    print(f'\nAn original and redo plate should be in the same pool, but that was not the case with {name[0]} and {p}.  Aborting script. Please redo the pooling .xlsx file\n\n')
                    
                    sys.exit()
                    
                    
                    
                
                ### can I remove the redo plates before doing the duplicate test? Remove if they match the pool and index of oringla plate, else keep
            
            
                

    return x_df
##########################
##########################



##########################
##########################
def assignPool(my_lib_df, my_pool_df):
    # remove rows were all columns are empty
    my_pool_df.dropna(axis=0, how='all', inplace=True)

    # abort script of any remaining cells are empty, e.g. didn't assign an Illumina index, pool#, or plate name
    if my_pool_df.isnull().values.any():
        print(
            "\n\nPooling Tool .xlsx file is missing a plate id, pool#, or illumina index. Aborting method\n\n")
        sys.exit()

    # count the number of lib plates in each pool
    # output is series sorted descending order, so index(0)
    # should be pool with most lib plates
    plates_per_pool = my_pool_df['Assigned_Pool'].value_counts()

    # The hamilton pool method only has space for 10 source plates, which means
    # pooling should be limited to 5 lib plates because there's potentially a
    # concentreated and dilute version of each lib plate.
    #
    # Flag if at least one pool as > 5 lib plates and ask if want to continue

    if plates_per_pool.index[0] > 5:
        print(
            "\n\nThe number of plates in at least one pool is >5. There probably will not be enough deck space.  Are you sure you want to continue?\n\n")

        val = input()

        if (val == 'Y' or val == 'y'):
            print("Ok, we'll keep going\n\n")

        elif (val == 'N' or val == 'n'):
            print(
                'Ok, aborting script.  Go back to .xslx pooling tool and adjust pool assignments\n\n')
            sys.exit()
        else:
            print("Sorry, you must choose 'Y' or 'N' next time. \n\nAborting\n\n")
            sys.exit()

    # my_pool_df.dropna(inplace=True)

    # make dict where key is source plate id and values is assigned pool from .xlsx pooling tool file
    pool_dict = dict(zip(my_pool_df.Plate, my_pool_df.Assigned_Pool))

    # add new column to my_passed_df showing assigned pool number for each plate
    my_lib_df['Pool_number'] = my_lib_df['Pool_source_plate'].map(pool_dict)
    
    # change pool number to -99 for libs that actually failed
    # this is a work around because clarity does not allow us to tail libs in plates
    # libs from pool number -99 are not processed by later pooling scripts
    my_lib_df['Pool_number'] = np.where(my_lib_df['Total_passed_attempts'] ==0,-99,my_lib_df['Pool_number'])
    
    my_lib_df['Pool_number'] = my_lib_df['Pool_number'].astype('Int64')

    # Destination_Tube_Name, Destination_Tube_Name

    my_lib_df['Destination_Tube_Name'] = 'Pool_' + \
        my_lib_df['Pool_number'].astype(str)

    my_lib_df['Destination_Tube_Barcode'] = 'Pool_' + \
        my_lib_df['Pool_number'].astype(str)

    # my_lib_df['Destination_Tube_Barcode'] = my_lib_df['Destination_Tube_Name']

    # sumarize df by pool number, index set, and source plates
    # This df will be searched for pools with duplicate index sets
    x_df = my_lib_df.groupby(['Pool_number', 'Pool_Illumina_index_set', 'Pool_source_plate'])[
        'Pool_Illumina_index'].agg('count').reset_index()

    # get rid of count column.  Don't need it.  Now have summry df of plates and index sets
    # in each pool number
    x_df.drop(columns=['Pool_Illumina_index'], inplace=True)

    # added this row in to test duplicated detection during testing.
    # x_df.loc[len(x_df.index)] = [3, 'Nextera_Index-5', 'XXXXX-X']
    
    # check if redo plates have same pool# and index as original. If so, remove before duplication test
    x_df = checkRedoPlates(x_df)

    # find any rows with where pool and index set are same.  This would be an
    # index colision within a pool
    dup_df = x_df[x_df.duplicated(
        ['Pool_number', 'Pool_Illumina_index_set'], keep=False)]

    # show duplicate indexes in pool if found, and abort script.
    if dup_df.shape[0] != 0:
        print('\n\n\n')
        print(dup_df)
        print('\n\n\n')
        print('\n\nAt least one pool has plates with same illumina index. Aborting')
        sys.exit()

    return my_lib_df, pool_dict
##########################
##########################


##########################
##########################
def getLibMass(my_passed_df):

    # # user provides min conc threshold for successful lib creation.
    # target_pool_mass = float(
    #     input("Enter the target pool MASS needed (default 2.75 pmol): ") or 2.75)

    # print('\n\n')

    # find number of libs in each plate
    pool_size_df = my_passed_df.groupby(['Pool_number'])[
        'Pool_source_well'].agg('count').reset_index()
    
    # remove pool number -99 since its only used to flag libs that failed
    pool_size_df.drop(pool_size_df.loc[pool_size_df['Pool_number']==-99].index, inplace=True)

    pool_size_df = pool_size_df.rename(
        columns={'Pool_source_well': 'Pool_size'})

    # determine how DNA mass required from each lib to reach target mass of pool
    # pool_size_df['Target_Mass_per_library'] = (target_pool_mass /
    #                                            pool_size_df['Pool_size']).round(3)

    pool_size_df['Target_Mass_per_library'] = (target_pool_mass /
                                               pool_size_df['Pool_size'])

    # make dict were key is plate ID and value is the number of libs on plate
    conc_dict = dict(zip(pool_size_df.Pool_number,
                         pool_size_df.Target_Mass_per_library))

    # add new column to my_passed_df showing assigned pool number for each plate
    my_passed_df['Pool_target_lib_mass_(pmol)'] = my_passed_df['Pool_number'].map(
        conc_dict)
    

    return my_passed_df, pool_size_df

##########################
##########################


##########################
##########################
def getLibVolumes(my_passed_df):

    # volume of concentrated lib need to meet per lib target mass
    my_passed_df['Pool_volume_concentrated_(uL)'] = (my_passed_df['Pool_target_lib_mass_(pmol)'] / (
        my_passed_df['Pool_nmole/L'] / 1000))

    
    if "Pool_dilution_factor" not in my_passed_df.columns:
        my_passed_df['Pool_dilution_factor'] = 1

    # volume of diluted lib need to meet per lib target mass
    my_passed_df['Pool_volume_diluted_(uL)'] = (my_passed_df['Pool_volume_concentrated_(uL)'] *
                                                my_passed_df['Pool_dilution_factor'])



    # determine if using concentrated or diluted lib plate as source for pooling based on
    # transfer volume need to reach per lib mass target

    my_passed_df['Pool_use_conc_or_dilut'] = np.where(
        my_passed_df['Pool_volume_concentrated_(uL)'] < min_tran_vol, 'dilute', 'concentrate')

    # use the concentrated source plate if the volume of diluted sample is >1.25 (actually variable concentrate_switch entered by user in .xslx) the max transfer volume from the dilute plate
    # this compromise leads to overshooting target mass transfer rather than undershoot if undershooting by more that 25%
    my_passed_df['Pool_use_conc_or_dilut'] = np.where(((
        my_passed_df['Pool_volume_concentrated_(uL)'] < min_tran_vol) & (my_passed_df['Pool_volume_diluted_(uL)'] > (max_dilut_vol*concentrate_switch))), 'concentrate', my_passed_df['Pool_use_conc_or_dilut'])

    my_passed_df['Pool_transfer_plate'] = np.where(my_passed_df['Pool_use_conc_or_dilut'] ==
                                                   'dilute', my_passed_df['Pool_source_plate']+'D', 'h' + my_passed_df['Pool_source_plate'])

    my_passed_df['Pool_transfer_volume_(uL)'] = np.where(my_passed_df['Pool_use_conc_or_dilut'] ==
                                                         'dilute', my_passed_df['Pool_volume_diluted_(uL)'], my_passed_df['Pool_volume_concentrated_(uL)'])

    # apply pipeting limits to transfer volumes based on limiations of liquid handler

    my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'] = my_passed_df['Pool_transfer_volume_(uL)']

    my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'] = np.where(((my_passed_df['Pool_use_conc_or_dilut'] == 'dilute') & (
        my_passed_df['Pool_transfer_volume_(uL)'] > max_dilut_vol)), max_dilut_vol, my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'])

    my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'] = np.where(((my_passed_df['Pool_use_conc_or_dilut'] == 'concentrate') & (
        my_passed_df['Pool_transfer_volume_(uL)'] > max_conc_vol)), max_conc_vol, my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'])

    my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'] = np.where(my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'] <
                                                                min_tran_vol, min_tran_vol, my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'])

    my_passed_df['Pool_target_lib_mass_(pmol)'] = my_passed_df['Pool_target_lib_mass_(pmol)'].round(
        4)
    my_passed_df['Pool_volume_concentrated_(uL)'] = my_passed_df['Pool_volume_concentrated_(uL)'].round(
        1)
    my_passed_df['Pool_volume_diluted_(uL)'] = my_passed_df['Pool_volume_diluted_(uL)'].round(
        1)
    my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'] = my_passed_df['Pool_ACTUAL_transfer_volume_(uL)'].round(
        1)

    my_passed_df['Pool_transfer_volume_(uL)'] = my_passed_df['Pool_transfer_volume_(uL)'].round(1)
    
    
    # replace null values associatd with failed libs with ""
    my_passed_df['Pool_volume_concentrated_(uL)'] = my_passed_df['Pool_volume_concentrated_(uL)'].fillna('')
    
    # replace null values associatd with failed libs with ""
    my_passed_df['Pool_target_lib_mass_(pmol)'] = my_passed_df['Pool_target_lib_mass_(pmol)'].fillna('')
    
    # replace concentrate/dilute plate assignment for failed libs
    my_passed_df['Pool_use_conc_or_dilut']=np.where(my_passed_df['Total_passed_attempts']==0,"",my_passed_df['Pool_use_conc_or_dilut'])

    return my_passed_df
##########################
##########################


##########################
##########################
def fillPoolPrepSheet(pool_file, pool_dict, pool_size_df):

    # make dict where key is pool# and value is number of libs in that pool
    pool_count_dict = dict(zip(pool_size_df.Pool_number,
                               pool_size_df.Pool_size))

    # make dict were key is plate ID and value is the number of libs on plate
    new_conc_dict = dict(zip(pool_size_df.Pool_number,
                         pool_size_df.Target_Mass_per_library))


    # import library creation .xls file downloaded from clarity queue
    clarity_df = pd.read_excel(
        pool_file, header=2, usecols="A:O")
    
    # create df that will eventually be used to fill in pool prep sheet section
    # asking for pool number, total pool %, and number libs in pool
    # find number of libs in each pool using the clarity pool prep sheet df
    pool_sum_df = clarity_df.groupby(['Pool Number'])[
        'Library Name'].agg('count').reset_index()
    
    pool_sum_df = pool_sum_df.rename(
        columns={'Library Name': 'Pool_size'})
    
    # add total pool % as fixed value
    pool_sum_df['Total Pool Percentage (%)'] = 100.0

    # remove columns that aren't needed to fill in pool prep .xls sheet
    pool_sum_df = pool_sum_df[[
        'Pool Number', 'Total Pool Percentage (%)', 'Pool_size']]

    
    # set pool number to nan so that it can be replace be real pool number
    # this is will be important once LIMS allows pools of SPS/CE to be reassigned
    # like we do with SIP
    clarity_df['Pool Number'] = np.nan

    # # add new column showing assigned pool number for each plate using pool dict to map
    # # pool # based on plate name
    # clarity_df['Pool Number'] = clarity_df['Plate Map'].map(pool_dict)
    
    ## this is difference between SPS and SIP pooling sheets
    # add new column showing assigned pool number for each plate using pool dict to map
    # pool # based on plate name
    clarity_df['Pool Number'] = clarity_df['Container ID'].map(pool_dict)
    
    clarity_df.dropna(subset=['Library Name'],inplace=True)

    if clarity_df['Pool Number'].isnull().values.any():
        print(
            "\n\nCould not assign pool number to one or more plates. Aborting method\n\n")
        sys.exit()

    # missing_plate = set(list(pool_dict.keys())) - \
    #     set(clarity_df['Plate Map'].unique().tolist())
    
    ## this version for identifying missing plates is needed for SPS instead of SIP
    missing_plate = set(list(pool_dict.keys())) - \
        set(clarity_df['Container ID'].unique().tolist())
        
    # remove and redo plates, i.e. plate names ending in ".#", e.g. 27-XXXXX.2
    # these redo plate names are not in the poolingprep file for SPS, which is
    # different than SIP.  They would appear to be missing plates if not removed    
    redo_plate_names = []   
    for p in missing_plate:
        
        if re.search("\.\d$",p):
            
            redo_plate_names.append(p)
        
    # remove all redo plates from set of missing plate names
    # if the length of missing plate set is >0, then abort the program
    if len((missing_plate-set(redo_plate_names))) > 0:
        print(
            f"\n\n Missing plates in pool assignment {missing_plate}. Aborting method\n\n")
        sys.exit()

    # # update degree of pooling column... but might not want to do
    # # this because DOP doesn't match pool size in file that comes from clarity
    # clarity_df['DOP'] = clarity_df['Pool Number'].map(pool_count_dict)

    # for clarity, calculate % of pool account for by each library in that pool
    clarity_df['Library Percentage with SOF (%)'] = (100 * (1 /
                                                            clarity_df['Pool Number'].map(pool_count_dict))).round(4)

    clarity_df['Library Molarity (pm)'] = clarity_df['Pool Number'].map(
        new_conc_dict).round(4)

    # make a copy of the pool creattion file, and fill new values into this copy
    copy_pool_file = 'autofilled_'+pool_file
    shutil.copyfile(pool_file, copy_pool_file)

    # load pool creation xlsx and delete existing values in first 2000 rows
    wb = openpyxl.load_workbook(copy_pool_file)

    # sh1 = wb['Sheet1']

    # newer format of pool prep file has different tab name
    sh1 = wb['Pooling Prep']

    sh1.delete_rows(4, 5000)

    # loop through clarity_df and write info into .xls file
    for row_num, row in clarity_df.iterrows():
        for col_num, col in enumerate(row):
            cellref = sh1.cell(row=(4+row_num), column=(1+col_num))
            cellref.value = col

    # loop through pool_size_dfdf and write info into .xls file
    for row_num, row in pool_sum_df.iterrows():
        for col_num, col in enumerate(row):
            # s.write(3+row_num, 16+col_num, col)
            cellref = sh1.cell(row=(4+row_num), column=(17+col_num))
            cellref.value = col

    total_lib_percent = 100 * pool_sum_df.shape[0]

    avg_lib_percent = 100

    # s.write(1, 18, total_lib_percent)

    cellref = sh1.cell(row=2, column=19)
    cellref.value = total_lib_percent

    # s.write(1, 19, avg_lib_percent)

    cellref = sh1.cell(row=2, column=20)
    cellref.value = avg_lib_percent

    # s.write(1, 16, "Done")

    cellref = sh1.cell(row=2, column=17)
    cellref.value = "Done"
    
    cellref = sh1.cell(row=2, column=18)
    cellref.value = 100

    wb.save(copy_pool_file)

    wb.close()
    return
##########################
##########################


##########################
# MAIN PROGRAM
##########################

###########################
# set up folder organiztion
###########################
ASSIGN_DIR = Path.cwd()

POOLING_DIR = ASSIGN_DIR.parent

PROJECT_DIR = POOLING_DIR.parent

ARCHIV_DIR = PROJECT_DIR / "archived_files"

FINISH_DIR = POOLING_DIR / "C_finish_pooling"

FINISH_DIR.mkdir(parents=True, exist_ok=True)


# get current working directory and its parent directory
crnt_dir = os.getcwd()
prnt_dir = os.path.dirname(crnt_dir)
prjct_dir = os.path.dirname(prnt_dir)



# get name of Clarity pool prep file, and abort script
# if 0 or >1 pool prep files are found
pool_file = findPoolPrepFile(crnt_dir)


# create df from project_summary.csv file
lib_df = pd.read_csv(PROJECT_DIR / 'project_summary.csv', header=0)

# read in assigned pool numbers from .xlsx pooling tool
pool_df = pd.read_excel('assign_pool_number_sheet.xlsx', header=1, engine=("openpyxl"), usecols=['Plate', 'Assigned_Pool', 'Index'],
                        converters={'Plate barcode': str, 'Assigned_Pool': int})


# get target mass and liquid handling constansts from .xlsx pooling tool
target_pool_mass, max_conc_vol, max_dilut_vol, min_tran_vol, concentrate_switch = getConstantsFromXlsx()

# updated lib_info_submitted df with manually assigned pool numbers
# also return dict where plate id is key and pool # is value
# and dict wher pool# is key and # of libs in pool is value
lib_df, pool_dict = assignPool(lib_df, pool_df)

# determine the per lib mass necessary for each pool
lib_df, pool_size_df = getLibMass(lib_df)

# determine the volume of each lib necessary to reach target mass.  Also determine if using concentrated or diluted lib plate
lib_df = getLibVolumes(lib_df)


# fill pool# into pool prep .xls file downloaded from clarity
fillPoolPrepSheet(pool_file, pool_dict, pool_size_df)


# lib_df.to_csv('review_pool.csv', index=False, header=True)


# get current date and time, will add to archive database file name
date = datetime.now().strftime("%Y_%m_%d-Time%H-%M-%S")


Path(PROJECT_DIR /
     "project_summary.csv").rename(ARCHIV_DIR / f"archive_project_summary_{date}.csv")
# Path(ARCHIV_DIR / f"archive_lib_info_submitted_to_clarity_{date}.csv").touch()

# create updated library info file
lib_df.to_csv(PROJECT_DIR / 'project_summary.csv', index=False)
